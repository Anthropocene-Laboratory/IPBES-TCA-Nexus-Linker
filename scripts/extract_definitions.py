"""Regenerate the application's reference data from the source workbook.

The application ships two read-only datasets, extracted verbatim from the
published IPBES assessments:

    src/data/tca_actions.json    22 transformative-change actions (TCA, chapter 5)
    src/data/nexus_options.json  71 response options (Nexus, chapter 5)

Both are derived from data/source/TCA and Nexus Definitions.xlsx. This script is
the single source of truth for that derivation: running it must reproduce the
committed JSON files byte for byte.

Usage:
    pip install -r scripts/requirements.txt
    python scripts/extract_definitions.py

Options:
    --check   Do not write; exit non-zero if the committed files are stale.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = ROOT / "data" / "source" / "TCA and Nexus Definitions.xlsx"
TCA_OUT = ROOT / "src" / "data" / "tca_actions.json"
NEXUS_OUT = ROOT / "src" / "data" / "nexus_options.json"

# Worksheet layout, 1-indexed columns, row 1 holds the headers.
TCA_SHEET = "TCA_Actions_Ch5"          # ID | Strategy | Action | Literal definition
NEXUS_SHEET = "Nexus_Response_Options"  # ID | Category | Code | Response option | Textual definition


def cell(value) -> str:
    """Normalise a cell to a stripped string; empty for blanks."""
    return "" if value is None else str(value).strip()


def read_tca_actions(workbook) -> list[dict]:
    """One record per action. Rows without an action title are ignored."""
    sheet = workbook[TCA_SHEET]
    records = []
    for row in range(2, sheet.max_row + 1):
        action = sheet.cell(row, 3).value
        if not action:
            continue
        records.append(
            {
                "id": cell(sheet.cell(row, 1).value),
                "strategy": cell(sheet.cell(row, 2).value),
                "action": cell(action),
                "definition": cell(sheet.cell(row, 4).value),
            }
        )
    return records


def read_nexus_options(workbook) -> list[dict]:
    """One record per response option, keyed by its published code (e.g. 'B01')."""
    sheet = workbook[NEXUS_SHEET]
    records = []
    for row in range(2, sheet.max_row + 1):
        title = sheet.cell(row, 4).value
        if not title:
            continue
        records.append(
            {
                "id": cell(sheet.cell(row, 3).value),
                "num": sheet.cell(row, 1).value,
                "category": cell(sheet.cell(row, 2).value),
                "title": cell(title),
                "definition": cell(sheet.cell(row, 5).value),
            }
        )
    return records


def serialise(records: list[dict]) -> str:
    return json.dumps(records, ensure_ascii=False, indent=2)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--check",
        action="store_true",
        help="verify the committed files match the workbook instead of rewriting them",
    )
    args = parser.parse_args()

    if not WORKBOOK.exists():
        print(f"Source workbook not found: {WORKBOOK}", file=sys.stderr)
        return 2

    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True)
    outputs = {
        TCA_OUT: read_tca_actions(workbook),
        NEXUS_OUT: read_nexus_options(workbook),
    }

    stale = []
    for path, records in outputs.items():
        payload = serialise(records)
        if args.check:
            current = path.read_text(encoding="utf-8") if path.exists() else None
            if current != payload:
                stale.append(path)
        else:
            path.write_text(payload, encoding="utf-8")
        print(f"{len(records):3d} records  {path.relative_to(ROOT).as_posix()}")

    if args.check and stale:
        for path in stale:
            print(f"STALE: {path.relative_to(ROOT).as_posix()}", file=sys.stderr)
        return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
